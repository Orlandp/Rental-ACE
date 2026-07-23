from flask import Blueprint, request, jsonify, session, send_file
from database import get_db
from routes.decorators import role_required
from datetime import datetime
import io

reports_bp = Blueprint('reports', __name__)


def get_report_data(month, year):
    conn = get_db()

    units = conn.execute('SELECT * FROM units').fetchall()
    units = [dict(u) for u in units]

    payments = conn.execute('''
        SELECT p.*, u.full_name AS tenant_name, un.unit_number
        FROM payments p
        JOIN users u ON p.tenant_id = u.user_id
        JOIN units un ON p.unit_id = un.unit_id
        WHERE p.month LIKE ?
    ''', (f'{month} {year}',)).fetchall()
    payments = [dict(p) for p in payments]

    expenses = conn.execute('''
        SELECT * FROM expenses WHERE expense_date LIKE ?
    ''', (f'{year}-%',)).fetchall()
    expenses = [dict(e) for e in expenses]

    conn.close()

    expected_income = sum(u['rent_amount'] for u in units)
    collected_income = sum(p['amount'] for p in payments if p['status'] == 'paid')
    outstanding = expected_income - collected_income
    total_expenses = sum(e['amount'] for e in expenses)
    net_income = collected_income - total_expenses

    return {
        'units': units,
        'payments': payments,
        'expenses': expenses,
        'expected_income': expected_income,
        'collected_income': collected_income,
        'outstanding': outstanding,
        'total_expenses': total_expenses,
        'net_income': net_income,
    }


@reports_bp.route('/api/reports/pdf', methods=['GET'])
@role_required('admin', 'landlord')
def download_pdf():
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    month = request.args.get('month', datetime.now().strftime('%B'))
    year = request.args.get('year', datetime.now().strftime('%Y'))

    data = get_report_data(month, year)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Ace Apartments — Financial Report", styles['Title']))
    elements.append(Paragraph(f"{month} {year}", styles['Normal']))
    elements.append(Spacer(1, 20))

    summary_data = [
        ['Expected Income', f"Ksh {data['expected_income']:,.0f}"],
        ['Collected Income', f"Ksh {data['collected_income']:,.0f}"],
        ['Outstanding', f"Ksh {data['outstanding']:,.0f}"],
        ['Total Expenses', f"Ksh {data['total_expenses']:,.0f}"],
        ['Net Income', f"Ksh {data['net_income']:,.0f}"],
    ]
    summary_table = Table(summary_data, colWidths=[200, 200])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a7a4a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 24))

    elements.append(Paragraph("Payments", styles['Heading2']))
    payment_rows = [['Unit', 'Tenant', 'Amount', 'Status', 'M-Pesa Code']]
    for p in data['payments']:
        payment_rows.append([
            f"H{p['unit_number']}",
            p['tenant_name'],
            f"Ksh {p['amount']:,.0f}",
            p['status'].upper(),
            p.get('mpesa_code') or '—',
        ])
    if len(payment_rows) > 1:
        payment_table = Table(payment_rows, colWidths=[50, 130, 90, 70, 100])
        payment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e8f5ee')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(payment_table)
    else:
        elements.append(Paragraph("No payments recorded for this period.", styles['Normal']))

    elements.append(Spacer(1, 24))
    elements.append(Paragraph("Expenses", styles['Heading2']))
    expense_rows = [['Category', 'Description', 'Amount', 'Date']]
    for e in data['expenses']:
        expense_rows.append([
            e['category'], e['description'], f"Ksh {e['amount']:,.0f}", e['expense_date'],
        ])
    if len(expense_rows) > 1:
        expense_table = Table(expense_rows, colWidths=[90, 180, 90, 80])
        expense_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fdecea')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(expense_table)
    else:
        elements.append(Paragraph("No expenses recorded for this period.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'rental-ace-report-{month}-{year}.pdf'
    )


@reports_bp.route('/api/reports/excel', methods=['GET'])
@role_required('admin', 'landlord')
def download_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    month = request.args.get('month', datetime.now().strftime('%B'))
    year = request.args.get('year', datetime.now().strftime('%Y'))

    data = get_report_data(month, year)

    wb = Workbook()

    summary_sheet = wb.active
    summary_sheet.title = 'Summary'
    header_fill = PatternFill(start_color='1a7a4a', end_color='1a7a4a', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    summary_sheet.append(['Rental-ACE Financial Report', f'{month} {year}'])
    summary_sheet.append([])
    summary_rows = [
        ('Expected Income', data['expected_income']),
        ('Collected Income', data['collected_income']),
        ('Outstanding', data['outstanding']),
        ('Total Expenses', data['total_expenses']),
        ('Net Income', data['net_income']),
    ]
    for label, value in summary_rows:
        summary_sheet.append([label, value])

    payments_sheet = wb.create_sheet('Payments')
    payments_sheet.append(['Unit', 'Tenant', 'Amount', 'Status', 'M-Pesa Code', 'Date'])
    for cell in payments_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for p in data['payments']:
        payments_sheet.append([
            p['unit_number'], p['tenant_name'], p['amount'],
            p['status'], p.get('mpesa_code') or '—', p['payment_date'],
        ])

    expenses_sheet = wb.create_sheet('Expenses')
    expenses_sheet.append(['Category', 'Description', 'Amount', 'Date'])
    for cell in expenses_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for e in data['expenses']:
        expenses_sheet.append([e['category'], e['description'], e['amount'], e['expense_date']])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'rental-ace-report-{month}-{year}.xlsx'
    )